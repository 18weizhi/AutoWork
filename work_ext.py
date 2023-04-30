import pyautogui as pag
import pyperclip
import time
import openpyxl

def mainwork():
    i = 2       #配置文件内容是从第2行开始
    while i <= sheet.max_row:
        cmdvalue = sheet.cell(row=i,column=1).value
        if cmdvalue == 1:                               #单击图片
            img = sheet.cell(row=i,column=2).value
            retry = 1
            if type(sheet.cell(row=i,column=4).value) != type(None):
                retry = sheet.cell(row=i,column=4).value
            mouseclick(1,'left',img,retry)
            print('单击左键1次')
            time.sleep(1)
        elif cmdvalue == 2:                             #双击
            img = sheet.cell(row=i,column=2).value
            retry = 1
            if sheet.cell(row=i,column=4).value != None:
                retry = sheet.cell(row=i,column=4).value
            mouseclick(2,'left',img,retry)
            print('单击左键2次')
            time.sleep(1)
        elif cmdvalue == 3:                             #右键
            img = sheet.cell(row=i,column=2).value
            retry = 1
            if sheet.cell(row=i,column=4).value != None:
                retry = sheet.cell(row=i,column=4).value
            mouseclick(1,'left',img,retry)
            print('单击右键1次')
            time.sleep(1)
        elif cmdvalue == 4:                             #输入
            inputvalue = sheet.cell(row=i,column=2).value
            pyperclip.copy(inputvalue)
            pag.hotkey('ctrl','v')
            print('输入' + inputvalue)
            time.sleep(0.5)
        elif cmdvalue == 5:                             #等待
            sleeptime = sheet.cell(row=i,column=2).value
            time.sleep(sleeptime)
            print('等待%s秒' % sleeptime)
        elif cmdvalue == 6:                             #滚轮
            scroll = sheet.cell(row=i,column=2).value
            pag.scroll(scroll)
            print('滚动滑轮')
        elif cmdvalue == 7:                             #单击坐标
            x = sheet.cell(row=i,column=2).value
            y = sheet.cell(row=i,column=3).value
            retry = 1
            if sheet.cell(row=i,column=4).value != None:
                retry = sheet.cell(row=i,column=4).value
            onceclick(x,y,retry)
            print('单击坐标%s,%s' % (x,y))
            time.sleep(1)
        elif cmdvalue == 8:                             #回车
            pag.press('enter')

        i+= 1

def datacheck():
    checkresult = True
    if sheet.max_row < 2:
        print('没有数据')
        checkresult = False
    i = 2
    while i <= sheet.max_row:
        cmdvalue = sheet.cell(row=i,column=1).value
        if type(cmdvalue) != type(1) or (cmdvalue != 1 and cmdvalue != 2 and cmdvalue != 3 and cmdvalue != 4 and cmdvalue != 5 and cmdvalue != 6 and cmdvalue != 7 and cmdvalue != 8):          #判断第一行第一列是否为数字并判断是否在1-6范围内
            print('第%s行，1列有问题' % i)
            checkresult = False
        if cmdvalue == 1:
            if type(sheet.cell(row=i,column=2).value) != type('1'):
                print('第%s行，2列有问题1' % i)
                checkresult = False
        elif cmdvalue == 2:
            if type(sheet.cell(row=i,column=2).value) != type('1'):
                print('第%s行，2列有问题2' % i)
                checkresult = False
        elif cmdvalue == 3:
            if type(sheet.cell(row=i,column=2).value) != type('1'):
                print('第%s行，2列有问题3' % i)
                checkresult = False
        elif cmdvalue == 4:
            if type(sheet.cell(row=i,column=2).value) != type('1'):
                print('第%s行，2列有问题4' % i)
                checkresult = False
        elif cmdvalue == 5:
            if type(sheet.cell(row=i,column=2).value) != type(1):
                print('第%s行，2列有问题5' % i)
                checkresult = False
        elif cmdvalue == 6:
            if type(sheet.cell(row=i,column=2).value) != type(1):
                print('第%s行，2列有问题6' % i)
                checkresult = False
        elif cmdvalue == 7:
            if type(sheet.cell(row=i,column=2).value) != type(1) or type(sheet.cell(row=i,column=3).value) != type(1):
                print('第%s行，2列或3列有问题7')
                chechresult = False
        elif cmdvalue == 8:
            if type(sheet.cell(row=i, column=2).value) != None:
                print('第%s行，2列有问题8' % i)
                checkresult = False
        i += 1
        return checkresult

def mouseclick(clicktimes,LorR,img,retry):
    if retry == 1:
        while True:
            location = pag.locateCenterOnScreen(img,confidence=0.9)
            if location is not None:
                pag.click(location.x,location.y,clicks=clicktimes,interval=0.2,duration=0.2,button=LorR)
                break
            print('未找到图片')
            time.sleep(0.5)
    elif retry == -1:
        while True:
            location = pag.locateCenterOnScreen(img, confidence=0.9)
            if location is not None:
                pag.click(location.x, location.y, clicks=clicktimes, interval=0.2, duration=0.2, button=LorR)
                print('重复')
                time.sleep(0.5)
    elif retry >= 2:
        while retry:
            location = pag.locateCenterOnScreen(img, confidence=0.9)
            if location is not None:
                pag.click(location.x, location.y, clicks=clicktimes, interval=0.2, duration=0.2, button=LorR)
                print('重复')
                time.sleep(0.5)
            retry -= 1
        

def onceclick(x,y,retry):
    if retry == 1:
        pag.click(x,y,button='left')
        time.sleep(0.5)
    elif retry == -1:
        while True:
            pag.click(x,y,button='left')
            time.sleep(1)
            print('重复')
    elif retry >= 2:
        while True:
            retry -= 1
            pag.click(x,y,button='left')
            time.sleep(1)
            print('重复')

if __name__ == '__main__':
    filename = 'cmd.xlsx'           #获取配置文件名字
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    result = datacheck()
    if result:
        key = input('请输入数字1或2{1代表执行1次，2代表无限循环}\n')
        if key == '1':
            mainwork()
        elif key == '2':
            while True:
                mainwork()
                time.sleep(0.5)
        else:
            print('请正确输入数字1或2')
    else:
        print("数据检测错误，请检查配置文件")
